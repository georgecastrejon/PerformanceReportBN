from docxtpl import DocxTemplate, InlineImage
import openpyxl

path = "input/data.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

tipoinforme = sheet_obj.cell(row = 5, column = 3).value
codigorq = sheet_obj.cell(row = 6, column = 3).value
nombrerq = sheet_obj.cell(row = 7, column = 3).value
ninforme = tipoinforme+' '+codigorq+ ' - '+nombrerq

doc = DocxTemplate("templates/template.docx")
context = {
    "titulo":sheet_obj.cell(row = 1, column = 3).value,
    "bannercli":InlineImage(doc,sheet_obj.cell(row = 2, column = 3).value),
    "piepagina":sheet_obj.cell(row = 3, column = 3).value,
    "cliente":sheet_obj.cell(row = 4, column = 3).value,
    "tipoinforme":tipoinforme,
    "codigorq":codigorq,
    "nombrerq":nombrerq,
    "fecharq":sheet_obj.cell(row = 8, column = 3).value,        
    "version":sheet_obj.cell(row = 9, column = 3).value,    
    "fecha":sheet_obj.cell(row = 10, column = 3).value,
    "responsable":sheet_obj.cell(row = 11, column = 3).value,
    "descripcioncambio":sheet_obj.cell(row = 12, column = 3).value,
    "termino1":sheet_obj.cell(row = 13, column = 3).value,
    "termino2":sheet_obj.cell(row = 14, column = 3).value,
    "termino3":sheet_obj.cell(row = 15, column = 3).value,
    "termino4":sheet_obj.cell(row = 16, column = 3).value,
    "termino5":sheet_obj.cell(row = 17, column = 3).value,
    "termino6":sheet_obj.cell(row = 18, column = 3).value,
    "termino7":sheet_obj.cell(row = 19, column = 3).value,
    "termino8":sheet_obj.cell(row = 20, column = 3).value,
    "termino9":sheet_obj.cell(row = 21, column = 3).value,
    "des_termino1":sheet_obj.cell(row = 22, column = 3).value,
    "des_termino2":sheet_obj.cell(row = 23, column = 3).value,
    "des_termino3":sheet_obj.cell(row = 24, column = 3).value,
    "des_termino4":sheet_obj.cell(row = 25, column = 3).value,
    "des_termino5":sheet_obj.cell(row = 26, column = 3).value,
    "des_termino6":sheet_obj.cell(row = 27, column = 3).value,
    "des_termino7":sheet_obj.cell(row = 28, column = 3).value,
    "des_termino8":sheet_obj.cell(row = 29, column = 3).value,
    "des_termino9":sheet_obj.cell(row = 30, column = 3).value,
    "introduccion":sheet_obj.cell(row = 31, column = 3).value,
    "consideraciones":sheet_obj.cell(row = 32, column = 3).value,
    "alcance":sheet_obj.cell(row = 33, column = 3).value,
    "estrategia":sheet_obj.cell(row = 34, column = 3).value,
    "conclusiones":sheet_obj.cell(row = 40, column = 3).value,
    "recomendaciones":sheet_obj.cell(row = 41, column = 3).value
}

doc.render(context)
doc.save("reports/"+ninforme+".docx")